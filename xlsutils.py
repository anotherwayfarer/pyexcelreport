#!/usr/bin/python
# -*- coding: utf-8 -*-

import os
from copy import copy

from openpyxl import Workbook, worksheet
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Font, Color, Alignment, PatternFill

def workbook_create():
    wb = Workbook()
    for i in wb.worksheets:
        wb.remove(i)
    return wb

def sheet_create(wb, main_sheet_name):
    wb.create_sheet( main_sheet_name )
    ws = wb.active

    return ws

def sheet_print_setup(ws, porientation, pwidth):
    'https://openpyxl.readthedocs.io/en/2.5/_modules/openpyxl/worksheet/page.html'

    ws.print_options.horizontalCentered = True
    ws.print_options.verticalCentered = False
    ws.print_options.headings = False
    ws.print_options.gridLines = False

    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0
    ws.page_margins.footer = 0

    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToHeight = False

    worksheet.Worksheet.set_printer_settings(ws, paper_size = 1, orientation=porientation)
    ws.page_setup.fitToWidth = pwidth
    if pwidth == 2:
        ws.print_options.horizontalCentered = False

def most_bottom_right_coords(start_row, start_col, end_row, end_col):
    'Returns the most right and the most bottom coordinates'
    new_end_row = start_row if (end_row is None) or (end_row < start_row) else end_row
    new_end_col = start_col if (end_col is None) or (end_col < start_col) else end_col
    return new_end_row, new_end_col

def apply_border( ws, start_row, start_col, end_row=None, end_col=None, border_style='thin' ):
    """
    """
    end_row, end_col = most_bottom_right_coords(start_row, start_col, end_row, end_col)
    new_border = Border(left=Side(style=border_style),
                        right=Side(style=border_style),
                        top=Side(style=border_style),
                        bottom=Side(style=border_style))

    for r in range( start_row, end_row + 1 ):
        for c in range( start_col, end_col + 1 ):
            ws.cell( row=r, column=c ).border = new_border

def apply_outline( ws, start_row, start_col, end_row=None, end_col=None, border_style='thin' ):
    """
    """
    def _apply_border( cl, side_name ):
        """Apply outline to the one cell
        """
        new_border = copy(cl.border)
        getattr(new_border, side_name).border_style = border_style
        cl.border = new_border

    end_row, end_col = most_bottom_right_coords(start_row, start_col, end_row, end_col)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            if (c not in [start_col, end_col]) and (r not in [start_row, end_row]):
                continue

            if c == start_col: _apply_border(ws.cell(row=r, column=start_col), 'left')
            if c == end_col:   _apply_border(ws.cell(row=r, column=end_col), 'right')
            if r == start_row: _apply_border(ws.cell(row=start_row, column=c), 'top')
            if r == end_row:   _apply_border(ws.cell(row=end_row, column=c), 'bottom')

def apply_font(ws, start_row, start_col, end_row=None, end_col=None, \
            name='Calibri', size=11, bold=False, italic=False, underline='none', \
            vertAlign='baseline', strike=False, color='FF000000'):
    """https://openpyxl.readthedocs.io/en/2.5/styles.html
    """
    end_row, end_col = most_bottom_right_coords(start_row, start_col, end_row, end_col)
    new_font = Font(name=name, size=size, bold=bold, italic=italic, underline=underline, \
               vertAlign=vertAlign, strike=strike, color=color)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).font = new_font

def apply_alignment(ws, start_row, start_col, end_row=None, end_col=None, \
            horizontal='center', vertical='center', textRotation=None, wrapText=True, \
            shrinkToFit=True):
    """https://openpyxl.readthedocs.io/en/2.5/_modules/openpyxl/styles/alignment.html
    """
    end_row, end_col = most_bottom_right_coords(start_row, start_col, end_row, end_col)
    new_align = Alignment(horizontal=horizontal, vertical=vertical, textRotation=textRotation, \
                    wrapText=wrapText, shrinkToFit=shrinkToFit)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).alignment = new_align

def apply_fill(ws, start_row, start_col, end_row=None, end_col=None, \
            color='FFFFFF', fill_type='solid'):
    """Fills the cell background with color
    """
    end_row, end_col = most_bottom_right_coords(start_row, start_col, end_row, end_col)
    new_fill = PatternFill(start_color=color, end_color=color, fill_type=fill_type)

    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).fill = new_fill

